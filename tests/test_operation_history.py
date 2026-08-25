from __future__ import annotations

import inspect
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from prime_ai_trader.app.controller import TradingController
from prime_ai_trader.core.models import Direction, Signal, SignalState
from prime_ai_trader.database.repository import Repository
from prime_ai_trader.history.export import export_operation_history
from prime_ai_trader.ui.dashboard import PrimeAITraderApp
from prime_ai_trader.ui.dialogs import DecisionHistoryDialog
from tests.helpers import synthetic_candles


class OperationHistoryRepositoryTests(unittest.TestCase):
    @staticmethod
    def _signal() -> Signal:
        return Signal(
            Direction.BUY, SignalState.CONFIRMED, 78,
            {"COMPRA": 0.66, "VENDA": 0.20, "AGUARDAR": 0.14},
            100.0, 1, ["Estrutura de alta", "Retomada compradora"],
            payout_percent=82, technical_score=75, buy_score=78,
            sell_score=41, buy_rule_points=74, sell_rule_points=22,
            buy_reasons=["Estrutura de alta", "Retomada compradora"],
            sell_reasons=["Correção temporária"],
            pullback_primary_direction="COMPRA",
            pullback_correction_direction="VENDA",
            pullback_phase="RETOMADA CONFIRMADA",
            pullback_depth_atr=0.72,
        )

    @classmethod
    def _snapshot(cls, *, signal_id: int | None = None, state: str = "SINAL CONFIRMADO") -> dict:
        signal = cls._signal()
        return {
            "created_at": "2026-08-25T12:00:00+00:00",
            "event_type": "SINAL CONFIRMADO" if signal_id else "ANÁLISE / AGUARDAR",
            "signal_id": signal_id,
            "market": "Criptomoedas", "symbol": "BTC/USDT", "timeframe": "1m",
            "horizon_minutes": 1, "platform": "VEX", "strategy": "crypto-test",
            "sensitivity": "RÁPIDO", "mode": "CONFIRMAÇÃO",
            "direction": "COMPRA" if signal_id else "AGUARDAR", "state": state,
            "score": 78, "payout_percent": 82, "stake_amount": 20.0,
            "pullback_state": "PULLBACK COMPRADOR CONFIRMADO",
            "market_regime": "TENDÊNCIA DE ALTA", "structure_event": "BOS",
            "reason_summary": "Retomada compradora", "technical_score": 75,
            "model_score": 66, "source_name": "Binance pública",
            "settings": {"market": "Criptomoedas", "timeframe": "1m", "mode": "CONFIRMAÇÃO",
                         "sensitivity": "RÁPIDO", "stake_amount": 20.0, "payout_percent": 82},
            "signal": {
                "technical_score": 75, "model_score": 66, "buy_score": 78, "sell_score": 41,
                "buy_rule_points": 74, "sell_rule_points": 22,
                "probabilities": signal.probabilities,
                "buy_reasons": signal.buy_reasons, "sell_reasons": signal.sell_reasons,
                "confluences": signal.confluences, "waiting_reasons": [], "blockers": [],
                "warnings": [], "independent_confirmations": ["momentum", "tendência"],
                "candlestick_patterns": ["MARUBOZU DE ALTA"], "reversal_reasons": [],
                "reversal_votes": 0, "confirmed_candle": True, "momentum_votes": 3,
                "pullback_primary_direction": "COMPRA", "pullback_correction_direction": "VENDA",
                "pullback_phase": "RETOMADA CONFIRMADA", "pullback_depth_atr": 0.72,
                "market_regime": "TENDÊNCIA DE ALTA", "structure_event": "BOS",
            },
            "indicators": {"rsi_14": 61.5, "ema_21": 99.7, "atr_14": 0.8},
            "features": {"reversal_pressure": 0.1, "micro_trend_atr": 0.35},
            "recent_candles": [{"open_time": "2026-08-25T11:59:00+00:00", "open": 99.8,
                                "high": 100.2, "low": 99.7, "close": 100.0,
                                "volume": 150, "taker_buy_volume": 90, "closed": True}],
            "news": [{"title": "Mercado normal", "published_at": "2026-08-25T11:45:00+00:00",
                      "source": "Reuters", "sentiment": "NEUTRA", "high_risk": False,
                      "url": "https://example.test"}],
            "economic_events": [{"currency": "USD", "event": "FOMC",
                                "scheduled_at": "2026-08-25T15:00:00+00:00", "impact": "HIGH"}],
            "visible_platform": {"price": 100.0, "remaining_seconds": 45, "payout_percent": 82},
        }

    def test_waiting_decision_persists_complete_settings_and_indicators(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            repository.record_decision(self._snapshot())
            rows = repository.decision_history()
            payload = json.loads(rows[0]["snapshot_json"])
        self.assertEqual(rows[0]["direction"], "AGUARDAR")
        self.assertEqual(payload["settings"]["sensitivity"], "RÁPIDO")
        self.assertEqual(payload["indicators"]["rsi_14"], 61.5)
        self.assertEqual(payload["signal"]["buy_score"], 78)

    def test_manual_result_creates_auditable_event_and_updates_linked_signal(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            signal_id = repository.save_signal(
                self._signal(), "Criptomoedas", "BTC/USDT", "1m", {}, "CONFIRMAÇÃO",
                platform="VEX", strategy="crypto-test", sensitivity="RÁPIDO", stake_amount=20,
            )
            repository.record_decision(self._snapshot(signal_id=signal_id))
            repository.record_manual_result(signal_id, "WIN", payout_percent=82, stake_amount=20)
            rows = repository.decision_history()
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["event_type"], "RESULTADO OBSERVADO")
        self.assertEqual(rows[0]["result_source"], "MANUAL")
        self.assertAlmostEqual(rows[0]["profit_loss"], 16.4)
        self.assertEqual(rows[1]["result"], "WIN")

    def test_inferred_result_is_explicitly_distinguished_from_platform_result(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            signal_id = repository.save_signal(
                self._signal(), "Criptomoedas", "BTC/USDT", "1m", {}, "CONFIRMAÇÃO",
            )
            repository.set_result(signal_id, 101, "WIN", result_source="INFERRED")
            row = repository.decision_history()[0]
        self.assertEqual(row["event_type"], "RESULTADO INFERIDO")
        self.assertIn("inferido", row["reason_summary"])

    def test_history_filters_by_asset_and_event_type(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            repository.record_decision(self._snapshot())
            eth = self._snapshot()
            eth["symbol"] = "ETH/USDT"
            eth["event_type"] = "CONFIGURAÇÃO ALTERADA"
            repository.record_decision(eth)
            rows = repository.decision_history(symbol="ETH/USDT", event_type="CONFIGURAÇÃO ALTERADA")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["symbol"], "ETH/USDT")

    def test_existing_database_receives_decision_history_without_erasing_signals(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "history.db"
            first = Repository(path)
            first.save_signal(self._signal(), "Criptomoedas", "BTC/USDT", "1m", {}, "CONFIRMAÇÃO")
            upgraded = Repository(path)
            upgraded.record_decision(self._snapshot())
            self.assertEqual(len(upgraded.recent()), 1)
            self.assertEqual(len(upgraded.decision_history()), 1)


class OperationHistoryControllerTests(unittest.TestCase):
    def test_live_analysis_logs_indicators_features_candles_and_all_settings(self) -> None:
        with tempfile.TemporaryDirectory() as temporary, patch.dict(os.environ, {"XDG_DATA_HOME": temporary}):
            controller = TradingController()
            controller.secrets["twelve_data_key"] = "SEGREDO_NUNCA_EXPORTAR"
            with patch.object(controller.binance, "fetch_candles", return_value=synthetic_candles(180)), patch.object(
                controller.news_provider, "fetch", return_value=[],
            ):
                controller.analyze()
            rows = controller.repository.decision_history()
            payload = json.loads(rows[0]["snapshot_json"])
            self.assertIn("rsi_14", payload["indicators"])
            self.assertIn("reversal_pressure", payload["features"])
            self.assertEqual(len(payload["recent_candles"]), 8)
            self.assertEqual(payload["settings"]["sensitivity"], controller.settings.sensitivity)
            self.assertNotIn("SEGREDO_NUNCA_EXPORTAR", rows[0]["snapshot_json"])

    def test_changing_sensitivity_is_recorded_as_configuration_event(self) -> None:
        with tempfile.TemporaryDirectory() as temporary, patch.dict(os.environ, {"XDG_DATA_HOME": temporary}):
            controller = TradingController()
            controller.settings.sensitivity = "RÁPIDO"
            controller.settings.payout_percent = 82
            controller.save_settings()
            row = controller.repository.decision_history()[0]
        self.assertEqual(row["event_type"], "CONFIGURAÇÃO ALTERADA")
        self.assertEqual(row["sensitivity"], "RÁPIDO")
        self.assertEqual(row["payout_percent"], 82)


class OperationHistoryExcelTests(unittest.TestCase):
    @staticmethod
    def _prepare(repository: Repository) -> None:
        signal = OperationHistoryRepositoryTests._signal()
        signal_id = repository.save_signal(
            signal, "Criptomoedas", "BTC/USDT", "1m", {"rsi_14": 61.5},
            "CONFIRMAÇÃO", platform="VEX", strategy="crypto-test", sensitivity="RÁPIDO",
            stake_amount=20,
        )
        repository.record_decision(OperationHistoryRepositoryTests._snapshot(signal_id=signal_id))
        repository.record_manual_result(signal_id, "WIN", payout_percent=82, stake_amount=20)

    def test_excel_contains_all_audit_sheets_and_correct_manual_finance(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            self._prepare(repository)
            output = export_operation_history(repository, Path(temporary) / "auditoria.xlsx")
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, [
                "Resumo", "Operações", "Decisões da IA", "Indicadores e features",
                "Configurações", "Velas e pullbacks", "Notícias e eventos",
            ])
            summary = {row[0]: row[1] for row in workbook["Resumo"].iter_rows(min_row=2, values_only=True)}
            self.assertEqual(summary["WIN observados"], 1)
            self.assertAlmostEqual(summary["Resultado financeiro observado (R$)"], 16.4)
            self.assertAlmostEqual(summary["Acerto mínimo para empatar (%)"], 54.9451, places=4)
            operations = list(workbook["Operações"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(operations[0][10], "COMPRA")
            self.assertEqual(operations[0][11], "WIN")
            self.assertEqual(operations[0][26], "COMPRA")
            self.assertEqual(operations[0][27], "VENDA")
            self.assertGreaterEqual(workbook["Indicadores e features"].max_row, 2)
            self.assertGreaterEqual(workbook["Velas e pullbacks"].max_row, 2)
            self.assertGreaterEqual(workbook["Notícias e eventos"].max_row, 3)
            workbook.close()

    def test_external_news_is_not_exported_as_excel_formula(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            snapshot = OperationHistoryRepositoryTests._snapshot()
            snapshot["news"][0]["title"] = '=HYPERLINK("https://malicious.test")'
            repository.record_decision(snapshot)
            output = export_operation_history(repository, Path(temporary) / "safe.xlsx")
            workbook = load_workbook(output)
            cell = workbook["Notícias e eventos"].cell(2, 6)
            self.assertEqual(cell.data_type, "s")
            self.assertTrue(str(cell.value).startswith("'="))
            workbook.close()

    def test_empty_history_still_exports_a_valid_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            output = export_operation_history(repository, Path(temporary) / "empty")
            workbook = load_workbook(output)
            self.assertEqual(output.suffix, ".xlsx")
            self.assertEqual(len(workbook.sheetnames), 7)
            self.assertEqual(workbook["Operações"].max_row, 1)
            workbook.close()

    def test_excel_does_not_mix_inferred_results_with_real_platform_accuracy(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            repository = Repository(Path(temporary) / "history.db")
            signal_id = repository.save_signal(
                OperationHistoryRepositoryTests._signal(), "Criptomoedas", "BTC/USDT", "1m", {},
                "CONFIRMAÇÃO", stake_amount=20,
            )
            repository.set_result(signal_id, 101, "WIN", result_source="INFERRED")
            workbook = load_workbook(export_operation_history(repository, Path(temporary) / "origin.xlsx"))
            summary = {row[0]: row[1] for row in workbook["Resumo"].iter_rows(min_row=2, values_only=True)}
            self.assertEqual(summary["Resultados realmente observados"], 0)
            self.assertEqual(summary["Resultados apenas inferidos"], 1)
            self.assertIsNone(summary["Acerto observado (%)"])
            workbook.close()


class OperationHistoryInterfaceTests(unittest.TestCase):
    def test_dashboard_exposes_complete_history_without_removing_existing_layout(self) -> None:
        source = inspect.getsource(PrimeAITraderApp._build_left)
        self.assertIn("HISTÓRICO COMPLETO / EXCEL", source)
        self.assertIn("self.open_decision_history", source)
        self.assertIn("self.open_manual_result", source)

    def test_excel_export_runs_in_background_and_uses_xlsx(self) -> None:
        source = inspect.getsource(DecisionHistoryDialog.export)
        self.assertIn(".xlsx", source)
        self.assertIn("threading.Thread", source)
        self.assertIn("self.parent._post_ui", source)

    def test_windows_installer_explicitly_includes_excel_dependency(self) -> None:
        root = Path(__file__).parents[1]
        requirements = (root / "requirements.txt").read_text(encoding="utf-8")
        spec = (root / "PrimeAITrader.spec").read_text(encoding="utf-8")
        self.assertIn("openpyxl", requirements)
        self.assertIn('"openpyxl"', spec)


if __name__ == "__main__":
    unittest.main()
