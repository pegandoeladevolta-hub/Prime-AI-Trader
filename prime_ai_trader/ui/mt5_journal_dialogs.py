from __future__ import annotations

from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .theme import COLORS


HEADER_FILL = PatternFill("solid", fgColor="102237")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")


def _window(parent, title: str, size: str) -> tk.Toplevel:
    window = tk.Toplevel(parent)
    window.title(title)
    window.geometry(size)
    window.configure(bg=COLORS["bg"])
    window.transient(parent)
    return window


def _money(value, currency: str) -> str:
    return f"{currency} {float(value or 0):,.2f}"


def _time(value) -> str:
    if not value:
        return "—"
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.astimezone().strftime("%d/%m/%Y %H:%M:%S") if parsed.tzinfo else parsed.strftime("%d/%m/%Y %H:%M:%S")
    except (TypeError, ValueError):
        return str(value)


class MT5PerformanceDialog:
    def __init__(self, parent, stats: dict, *, currency: str = "USD") -> None:
        self.window = _window(parent, "Desempenho MT5", "720x540")
        outer = ttk.Frame(self.window, style="Panel.TFrame", padding=20)
        outer.pack(fill="both", expand=True, padx=12, pady=12)
        ttk.Label(outer, text="DESEMPENHO REAL • MT5", style="Panel.TLabel",
                  font=("Segoe UI Semibold", 15)).pack(anchor="w")
        ttk.Label(
            outer,
            text="Calculado com as operações executadas e encerradas no MetaTrader 5. "
                 "Não usa payout nem valor de entrada de opções binárias.",
            style="Muted.TLabel", wraplength=650,
        ).pack(anchor="w", pady=(4, 14))

        total = int(stats.get("total") or 0)
        accuracy = stats.get("accuracy")
        pf = stats.get("profit_factor")
        avg_r = stats.get("average_r")
        summary = (
            f"Operações encerradas: {total}\n"
            f"WIN: {stats.get('wins', 0)}   LOSS: {stats.get('losses', 0)}   DRAW: {stats.get('draws', 0)}\n"
            f"Acerto direcional: {accuracy * 100:.2f}%" if accuracy is not None else
            f"Operações encerradas: {total}\nWIN: 0   LOSS: 0   DRAW: 0\nAcerto direcional: sem amostra"
        )
        summary += (
            f"\nTP: {stats.get('tp', 0)}   SL: {stats.get('sl', 0)}   Manual: {stats.get('manual', 0)}"
            f"\nLucro bruto: {_money(stats.get('gross_profit'), currency)}   "
            f"Perda bruta: {_money(stats.get('gross_loss'), currency)}"
            f"\nResultado líquido: {_money(stats.get('net_profit'), currency)}"
            f"\nProfit factor: {pf:.2f}" if pf is not None else
            f"\nTP: {stats.get('tp', 0)}   SL: {stats.get('sl', 0)}   Manual: {stats.get('manual', 0)}"
            f"\nLucro bruto: {_money(stats.get('gross_profit'), currency)}   "
            f"Perda bruta: {_money(stats.get('gross_loss'), currency)}"
            f"\nResultado líquido: {_money(stats.get('net_profit'), currency)}"
            "\nProfit factor: —"
        )
        summary += (
            f"\nExpectativa/operação: {_money(stats.get('expectancy_per_operation'), currency)}"
            f"\nR realizado médio: {avg_r:+.2f}R" if avg_r is not None else
            f"\nExpectativa/operação: {_money(stats.get('expectancy_per_operation'), currency)}"
            "\nR realizado médio: —"
        )
        ttk.Label(outer, text=summary, style="Panel.TLabel", font=("Segoe UI", 12),
                  justify="left").pack(anchor="w", pady=(8, 18))
        ttk.Label(
            outer,
            text="Fonte financeira: negócios do MT5 (profit + commission + swap + fee).",
            style="Muted.TLabel",
        ).pack(anchor="w")
        ttk.Button(outer, text="FECHAR", command=self.window.destroy).pack(anchor="e", pady=(18, 0))


class MT5TradeHistoryDialog:
    COLUMNS = (
        "opened", "closed", "symbol", "tf", "direction", "volume", "entry",
        "sl", "tp", "rr", "exit", "result", "pnl", "reason", "strategy",
    )

    def __init__(self, parent, journal, rows: list[dict], *, currency: str = "USD") -> None:
        self.parent = parent
        self.journal = journal
        self.rows = rows
        self.currency = currency
        self.window = _window(parent, "Histórico operacional MT5", "1400x760")
        self.window.minsize(1050, 620)
        outer = ttk.Frame(self.window, style="Panel.TFrame", padding=16)
        outer.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Label(outer, text="DIÁRIO OPERACIONAL • METATRADER 5", style="Panel.TLabel",
                  font=("Segoe UI Semibold", 15)).pack(anchor="w")
        ttk.Label(
            outer,
            text="Registra somente a operação atual: lote, entrada, Stop Loss, Take Profit, R:R, "
                 "saída, P/L e motivo do encerramento. Payout e stake não fazem parte deste diário.",
            style="Muted.TLabel", wraplength=1200,
        ).pack(anchor="w", pady=(4, 10))

        frame = ttk.Frame(outer, style="Panel.TFrame")
        frame.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(frame, columns=self.COLUMNS, show="headings", height=18)
        definitions = (
            ("opened", "ABERTURA", 140), ("closed", "ENCERRAMENTO", 140),
            ("symbol", "ATIVO", 80), ("tf", "TF", 45), ("direction", "DIREÇÃO", 75),
            ("volume", "LOTE", 65), ("entry", "ENTRADA", 85), ("sl", "SL", 85),
            ("tp", "TP", 85), ("rr", "R:R", 60), ("exit", "SAÍDA", 85),
            ("result", "RESULTADO", 75), ("pnl", f"P/L {currency}", 90),
            ("reason", "MOTIVO SAÍDA", 125), ("strategy", "ESTRATÉGIA", 180),
        )
        for key, title, width in definitions:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, minwidth=45, anchor="center")
        vertical = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        horizontal = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vertical.set, xscrollcommand=horizontal.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vertical.grid(row=0, column=1, sticky="ns")
        horizontal.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.detail = tk.Text(outer, height=7, bg=COLORS["card_alt"], fg=COLORS["text"],
                              relief="flat", wrap="word", font=("Consolas", 9), padx=10, pady=8)
        self.detail.pack(fill="x", pady=(10, 0))
        self.tree.bind("<<TreeviewSelect>>", self._show_detail)

        actions = ttk.Frame(outer, style="Panel.TFrame")
        actions.pack(fill="x", pady=(10, 0))
        ttk.Button(actions, text="ATUALIZAR", command=self.refresh).pack(side="left")
        ttk.Button(actions, text="EXPORTAR EXCEL (.XLSX)", style="Accent.TButton",
                   command=self.export).pack(side="left", padx=8)
        ttk.Button(actions, text="FECHAR", command=self.window.destroy).pack(side="right")
        self._populate()

    @staticmethod
    def _price(value) -> str:
        return f"{float(value):g}" if value not in (None, "") else "—"

    def _populate(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for row in self.rows:
            rr = f"1:{float(row.get('risk_reward')):.2f}" if row.get("risk_reward") else "—"
            pnl = _money(row.get("net_profit"), self.currency) if row.get("status") == "ENCERRADA" else "—"
            self.tree.insert("", "end", iid=str(row["id"]), values=(
                _time(row.get("opened_at")), _time(row.get("closed_at")), row.get("symbol"),
                row.get("timeframe"), row.get("direction"), f"{float(row.get('volume') or 0):g}",
                self._price(row.get("entry_price")), self._price(row.get("stop_loss")),
                self._price(row.get("take_profit")), rr, self._price(row.get("exit_price")),
                row.get("result"), pnl, row.get("exit_reason") or "—", row.get("strategy") or "—",
            ))
        if self.rows:
            self.tree.selection_set(str(self.rows[0]["id"]))
            self._show_detail()

    def _show_detail(self, _event=None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        row = next((item for item in self.rows if int(item["id"]) == int(selected[0])), None)
        if row is None:
            return
        text = (
            f"ATIVO: {row.get('symbol')}   DIREÇÃO: {row.get('direction')}   LOTE: {row.get('volume')}\n"
            f"ENTRADA: {self._price(row.get('entry_price'))}   SL: {self._price(row.get('stop_loss'))}   "
            f"TP: {self._price(row.get('take_profit'))}   R:R: {row.get('risk_reward') or '—'}\n"
            f"SAÍDA: {self._price(row.get('exit_price'))}   RESULTADO: {row.get('result')}   "
            f"P/L LÍQUIDO: {_money(row.get('net_profit'), self.currency)}   MOTIVO: {row.get('exit_reason') or '—'}\n"
            f"COMISSÃO: {row.get('commission') or 0}   SWAP: {row.get('swap') or 0}   TAXA: {row.get('fee') or 0}\n"
            f"ESTRATÉGIA: {row.get('strategy') or '—'}   PERFIL: {row.get('sensitivity') or '—'}   "
            f"MODO: {row.get('mode') or '—'}   GESTÃO: {row.get('management') or '—'}   SCORE: {row.get('score') or 0}"
        )
        self.detail.configure(state="normal")
        self.detail.delete("1.0", "end")
        self.detail.insert("end", text)
        self.detail.configure(state="disabled")

    def refresh(self) -> None:
        self.rows = self.journal.recent(2000)
        self._populate()

    def export(self) -> None:
        destination = filedialog.asksaveasfilename(
            parent=self.window, title="Exportar diário MT5", defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            initialfile=f"PrimeTrader-MT5-Historico-{datetime.now():%Y-%m-%d_%H-%M}.xlsx",
        )
        if not destination:
            return
        output = export_mt5_journal(self.journal, destination, currency=self.currency)
        messagebox.showinfo("Histórico MT5", f"Diário exportado com sucesso.\n\n{output}", parent=self.window)


def export_mt5_journal(journal, destination: str | Path, *, currency: str = "USD") -> Path:
    output = Path(destination)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    rows = list(reversed(journal.recent(100000)))
    stats = journal.statistics()
    book = Workbook()
    summary = book.active
    summary.title = "Resumo MT5"
    summary.append(["PRIME TRADER — DIÁRIO MT5", ""])
    summary.append(["Operações encerradas", stats.get("total", 0)])
    summary.append(["WIN", stats.get("wins", 0)])
    summary.append(["LOSS", stats.get("losses", 0)])
    summary.append(["DRAW", stats.get("draws", 0)])
    accuracy = stats.get("accuracy")
    summary.append(["Acerto direcional (%)", accuracy * 100 if accuracy is not None else None])
    summary.append([f"Resultado líquido ({currency})", stats.get("net_profit", 0)])
    summary.append(["Profit factor", stats.get("profit_factor")])
    summary.append(["R realizado médio", stats.get("average_r")])
    for cell in summary[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 24

    sheet = book.create_sheet("Operações MT5")
    headings = [
        "ID", "Abertura", "Encerramento", "Ativo", "Timeframe", "Direção", "Lote",
        "Preço entrada", "Stop Loss", "Take Profit", "R:R", "Preço saída", "R realizado",
        f"Lucro bruto ({currency})", f"Comissão ({currency})", f"Swap ({currency})",
        f"Taxa ({currency})", f"P/L líquido ({currency})", "Resultado", "Motivo saída",
        "Estratégia", "Perfil", "Modo", "Gestão", "Score", "Ticket posição", "Ticket negócio",
    ]
    sheet.append(headings)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([
            row.get("id"), _time(row.get("opened_at")), _time(row.get("closed_at")),
            row.get("symbol"), row.get("timeframe"), row.get("direction"), row.get("volume"),
            row.get("entry_price"), row.get("stop_loss"), row.get("take_profit"), row.get("risk_reward"),
            row.get("exit_price"), row.get("realized_r"), row.get("profit"), row.get("commission"),
            row.get("swap"), row.get("fee"), row.get("net_profit"), row.get("result"),
            row.get("exit_reason"), row.get("strategy"), row.get("sensitivity"), row.get("mode"),
            row.get("management"), row.get("score"), row.get("position_ticket"), row.get("deal_ticket"),
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.iter_cols(max_row=min(sheet.max_row, 40)), 1):
        width = max((len(str(cell.value or "")) for cell in column), default=10)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 11), 34)
    book.save(output)
    return output


__all__ = ["MT5PerformanceDialog", "MT5TradeHistoryDialog", "export_mt5_journal"]
