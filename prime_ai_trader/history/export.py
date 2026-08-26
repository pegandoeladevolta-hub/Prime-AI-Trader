from __future__ import annotations

import json
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..database.repository import Repository


HEADER_FILL = PatternFill("solid", fgColor="102237")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
WIN_FILL = PatternFill("solid", fgColor="E4F5E9")
LOSS_FILL = PatternFill("solid", fgColor="FBE8E8")
WAIT_FILL = PatternFill("solid", fgColor="FFF4D5")
MONEY_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def _safe(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, datetime):
        value = value.astimezone().strftime("%d/%m/%Y %H:%M:%S")
    elif isinstance(value, (list, tuple, dict)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))[:32_700]
    # Notícias e rótulos chegam de fontes externas; nunca viram fórmulas Excel.
    if text and text[0] in ("=", "+", "-", "@", "\t", "\r"):
        text = "'" + text
    return text


def _time(value: Any) -> str:
    if not value:
        return ""
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone()
        return parsed.strftime("%d/%m/%Y %H:%M:%S")
    except (TypeError, ValueError):
        return str(value)


def _decode(value: Any, default: Any) -> Any:
    if isinstance(value, type(default)):
        return value
    try:
        found = json.loads(value or "")
    except (TypeError, ValueError, json.JSONDecodeError):
        return default
    return found if isinstance(found, type(default)) else default


def _sheet(workbook: Workbook, title: str, headings: list[str]):
    sheet = workbook.create_sheet(title)
    sheet.append(headings)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 25
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return sheet


def _finish(sheet) -> None:
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.iter_cols(max_row=min(sheet.max_row, 35)), start=1):
        length = max((len(str(cell.value or "")) for cell in column), default=10)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(length + 2, 12), 54)


def _append(sheet, values: list[Any]) -> None:
    sheet.append([_safe(value) for value in values])


def export_operation_history(repository: Repository, destination: str | Path, *,
                             limit: int = 25_000) -> Path:
    """Exporta auditoria local real, sem credenciais e sem resultados inventados."""
    output = Path(destination)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    decisions = list(reversed(repository.decision_history(limit)))
    operations = list(reversed(repository.recent(limit)))
    decoded = {int(row["id"]): _decode(row.get("snapshot_json"), {}) for row in decisions}
    signal_decisions: dict[int, dict] = {}
    for row in decisions:
        signal_id = row.get("signal_id")
        if signal_id is not None and row.get("event_type") == "SINAL CONFIRMADO":
            signal_decisions[int(signal_id)] = decoded[int(row["id"])]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 43
    summary.append(["PRIME AI TRADER — AUDITORIA OPERACIONAL", ""])
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    summary.row_dimensions[1].height = 26
    observed = [row for row in operations if row.get("result_source") == "MANUAL"
                and row.get("result") in {"WIN", "LOSS", "DRAW"}]
    inferred = [row for row in operations if row.get("result_source") == "INFERRED"
                and row.get("result") in {"WIN", "LOSS", "DRAW"}]
    wins = sum(row.get("result") == "WIN" for row in observed)
    losses = sum(row.get("result") == "LOSS" for row in observed)
    draws = sum(row.get("result") == "DRAW" for row in observed)
    directional = wins + losses
    payouts = [float(row.get("payout_percent") or 0) for row in observed
               if row.get("result") in {"WIN", "LOSS"}]
    average_payout = sum(payouts) / len(payouts) if payouts else None
    overview = [
        ("Gerado em", _time(datetime.now(timezone.utc).isoformat())),
        ("Leituras e eventos exportados", len(decisions)),
        ("Sinais confirmados registrados", len(operations)),
        ("Resultados realmente observados", len(observed)),
        ("WIN observados", wins),
        ("LOSS observados", losses),
        ("DRAW observados", draws),
        ("Acerto observado (%)", round(wins / directional * 100, 4) if directional else None),
        ("Payout observado médio (%)", round(average_payout, 4) if average_payout else None),
        ("Acerto mínimo para empatar (%)",
         round(100 / (1 + average_payout / 100), 4) if average_payout else None),
        ("Resultado financeiro observado (R$)",
         round(sum(float(row.get("profit_loss") or 0) for row in observed), 2)),
        ("Resultados apenas inferidos", len(inferred)),
        ("Resultado financeiro apenas inferido (R$)",
         round(sum(float(row.get("profit_loss") or 0) for row in inferred), 2)),
        ("Critério de verdade", "Somente WIN/LOSS registrados manualmente na plataforma"),
        ("Importante", "Resultado inferido pelo preço público não comprova execução na plataforma"),
        ("Segurança", "Nenhuma chave de API, senha, cookie, saldo ou token é exportado"),
    ]
    for label, value in overview:
        _append(summary, [label, value])
    for row in (12, 14):
        summary.cell(row=row, column=2).number_format = MONEY_FORMAT

    operation_headings = [
        "ID sinal", "Data/hora", "Plataforma", "Mercado", "Ativo", "Timeframe",
        "Expiração (min)", "Sensibilidade", "Modo", "Estratégia", "Direção",
        "Resultado", "Origem do resultado", "Payout (%)", "Entrada (R$)",
        "Lucro/prejuízo (R$)", "Preço de entrada", "Preço de saída", "Score",
        "Score técnico", "Score IA", "Score compra", "Score venda", "Pontos compra",
        "Pontos venda", "Regime", "Tendência principal", "Direção da correção",
        "Fase do pullback", "Profundidade ATR", "Evento estrutural", "Risco reversão",
        "Votos reversão", "Motivos compra", "Motivos venda", "Motivos da entrada",
        "Stop técnico", "Alvo técnico", "Espaço R", "Modelo", "Resultado registrado em",
    ]
    sheet = _sheet(workbook, "Operações", operation_headings)
    for row in operations:
        detail = signal_decisions.get(int(row["id"]), {})
        signal = detail.get("signal", {})
        _append(sheet, [
            row["id"], _time(row.get("created_at")), row.get("platform"), row.get("market"),
            row.get("symbol"), row.get("timeframe"), row.get("horizon_minutes"),
            row.get("sensitivity"), row.get("mode"), row.get("strategy"), row.get("direction"),
            row.get("result") or "PENDENTE", row.get("result_source") if row.get("result") else "",
            row.get("payout_percent"), row.get("stake_amount"), row.get("profit_loss"),
            row.get("entry"), row.get("exit"), row.get("score"), signal.get("technical_score"),
            signal.get("model_score"), signal.get("buy_score"), signal.get("sell_score"),
            signal.get("buy_rule_points"), signal.get("sell_rule_points"),
            signal.get("market_regime"), signal.get("pullback_primary_direction"),
            signal.get("pullback_correction_direction"), signal.get("pullback_phase"),
            signal.get("pullback_depth_atr"), signal.get("structure_event"),
            signal.get("reversal_risk"), signal.get("reversal_votes"),
            " | ".join(signal.get("buy_reasons", [])),
            " | ".join(signal.get("sell_reasons", [])),
            " | ".join(_decode(row.get("confluences_json"), [])),
            row.get("technical_stop"), row.get("technical_target"), row.get("technical_room_ratio"),
            row.get("model_version"), _time(row.get("result_observed_at")),
        ])
        result = row.get("result")
        if result in {"WIN", "LOSS"}:
            sheet.cell(sheet.max_row, 12).fill = WIN_FILL if result == "WIN" else LOSS_FILL
        for column in (15, 16):
            sheet.cell(sheet.max_row, column).number_format = MONEY_FORMAT
    _finish(sheet)

    decision_headings = [
        "ID evento", "Data/hora", "Tipo", "ID sinal", "Plataforma", "Mercado", "Ativo",
        "Timeframe", "Expiração (min)", "Sensibilidade", "Modo", "Estratégia",
        "Direção final", "Estado", "Resultado vinculado", "Origem resultado", "Payout (%)",
        "Entrada (R$)", "Lucro/prejuízo (R$)", "Score combinado", "Score técnico",
        "Score IA", "Score compra", "Score venda", "Pontos compra", "Pontos venda",
        "Probabilidade compra", "Probabilidade venda", "Probabilidade aguardar",
        "Regime", "Viés timeframe superior", "Timeframe superior real",
        "Regime timeframe superior", "Candles timeframe superior",
        "Fonte timeframe superior", "Tendência do pullback",
        "Correção do pullback", "Fase do pullback", "Profundidade ATR", "Evento estrutural",
        "Setup", "Vela confirmada", "Momentum votos", "Confirmações independentes",
        "Votos reversão", "Motivos reversão", "Motivos compra", "Motivos venda",
        "Motivos espera", "Bloqueios", "Alertas", "Padrões de candle", "Fonte",
        "Atraso fonte (s)", "Preço plataforma", "Tempo plataforma (s)",
        "Payout plataforma (%)", "Resumo", "Notícias", "Eventos econômicos",
    ]
    sheet = _sheet(workbook, "Decisões da IA", decision_headings)
    for row in decisions:
        detail = decoded[int(row["id"])]
        signal = detail.get("signal", {})
        probabilities = signal.get("probabilities", {})
        visible = detail.get("visible_platform") or {}
        _append(sheet, [
            row["id"], _time(row.get("created_at")), row.get("event_type"), row.get("signal_id"),
            row.get("platform"), row.get("market"), row.get("symbol"), row.get("timeframe"),
            row.get("horizon_minutes"), row.get("sensitivity"), row.get("mode"), row.get("strategy"),
            row.get("direction"), row.get("state"), row.get("result"), row.get("result_source"),
            row.get("payout_percent"), row.get("stake_amount"), row.get("profit_loss"),
            row.get("score"), row.get("technical_score"), row.get("model_score"),
            signal.get("buy_score"), signal.get("sell_score"), signal.get("buy_rule_points"),
            signal.get("sell_rule_points"), probabilities.get("COMPRA"), probabilities.get("VENDA"),
            probabilities.get("AGUARDAR"), row.get("market_regime"),
            signal.get("higher_timeframe_bias"), signal.get("higher_timeframe_label"),
            signal.get("higher_timeframe_regime"), signal.get("higher_timeframe_candles"),
            signal.get("higher_timeframe_source"), signal.get("pullback_primary_direction"),
            signal.get("pullback_correction_direction"), signal.get("pullback_phase"),
            signal.get("pullback_depth_atr"), row.get("structure_event"), signal.get("setup_name"),
            signal.get("confirmed_candle"), signal.get("momentum_votes"),
            " | ".join(signal.get("independent_confirmations", [])), signal.get("reversal_votes"),
            " | ".join(signal.get("reversal_reasons", [])),
            " | ".join(signal.get("buy_reasons", [])), " | ".join(signal.get("sell_reasons", [])),
            " | ".join(signal.get("all_waiting_reasons") or signal.get("waiting_reasons", [])),
            " | ".join(signal.get("blockers", [])),
            " | ".join(signal.get("warnings", [])), " | ".join(signal.get("candlestick_patterns", [])),
            row.get("source_name"), detail.get("source_lag_seconds"), visible.get("price"),
            visible.get("remaining_seconds"), visible.get("payout_percent"),
            row.get("reason_summary"), len(detail.get("news", [])),
            len(detail.get("economic_events", [])),
        ])
        if "AGUARDAR" in str(row.get("event_type")):
            sheet.cell(sheet.max_row, 3).fill = WAIT_FILL
    _finish(sheet)

    indicator_keys = sorted({key for detail in decoded.values()
                             for key in detail.get("indicators", {})})
    feature_keys = sorted({key for detail in decoded.values()
                           for key in detail.get("features", {})})
    sheet = _sheet(workbook, "Indicadores e features", [
        "ID evento", "Data/hora", "Ativo", "Timeframe", "Direção", "Estado",
        *[f"ind.{key}" for key in indicator_keys],
        *[f"feat.{key}" for key in feature_keys],
    ])
    for row in decisions:
        detail = decoded[int(row["id"])]
        indicators = detail.get("indicators", {})
        features = detail.get("features", {})
        if not indicators and not features:
            continue
        _append(sheet, [row["id"], _time(row.get("created_at")), row.get("symbol"),
                        row.get("timeframe"), row.get("direction"), row.get("state"),
                        *[indicators.get(key) for key in indicator_keys],
                        *[features.get(key) for key in feature_keys]])
    _finish(sheet)

    config_keys = sorted({key for detail in decoded.values()
                          for key in detail.get("settings", {})})
    sheet = _sheet(workbook, "Configurações", [
        "ID evento", "Data/hora", "Tipo", "Ativo", *config_keys,
    ])
    previous_signature = ""
    for row in decisions:
        detail = decoded[int(row["id"])]
        settings = detail.get("settings", {})
        if not settings:
            continue
        signature = json.dumps(settings, ensure_ascii=False, sort_keys=True)
        if signature == previous_signature and row.get("event_type") != "CONFIGURAÇÃO ALTERADA":
            continue
        previous_signature = signature
        _append(sheet, [row["id"], _time(row.get("created_at")), row.get("event_type"),
                        row.get("symbol"), *[settings.get(key) for key in config_keys]])
    _finish(sheet)

    sheet = _sheet(workbook, "Velas e pullbacks", [
        "ID evento", "Ativo", "Timeframe", "Tendência principal", "Correção",
        "Fase", "Posição da vela", "Abertura da vela", "Open", "High", "Low",
        "Close", "Volume", "Taker compra", "Fechada", "Direção final", "Resultado",
    ])
    for row in decisions:
        detail = decoded[int(row["id"])]
        signal = detail.get("signal", {})
        if not (signal.get("pullback_phase") or signal.get("pullback_state")
                or row.get("event_type") == "SINAL CONFIRMADO"):
            continue
        candles = detail.get("recent_candles", [])
        for index, candle in enumerate(candles, 1):
            _append(sheet, [row["id"], row.get("symbol"), row.get("timeframe"),
                            signal.get("pullback_primary_direction"),
                            signal.get("pullback_correction_direction"),
                            signal.get("pullback_phase"), f"{index}/{len(candles)}",
                            _time(candle.get("open_time")), candle.get("open"),
                            candle.get("high"), candle.get("low"), candle.get("close"),
                            candle.get("volume"), candle.get("taker_buy_volume"),
                            candle.get("closed"), row.get("direction"), row.get("result")])
    _finish(sheet)

    sheet = _sheet(workbook, "Notícias e eventos", [
        "ID evento", "Data/hora análise", "Ativo", "Categoria", "Publicado/agendado",
        "Título/evento", "Fonte/moeda", "Sentimento/impacto", "Alto risco", "URL",
    ])
    seen: set[tuple[str, str, str]] = set()
    for row in decisions:
        detail = decoded[int(row["id"])]
        for item in detail.get("news", []):
            key = ("news", str(item.get("title")), str(item.get("published_at")))
            if key in seen:
                continue
            seen.add(key)
            _append(sheet, [row["id"], _time(row.get("created_at")), row.get("symbol"),
                            "NOTÍCIA", _time(item.get("published_at")), item.get("title"),
                            item.get("source"), item.get("sentiment"), item.get("high_risk"),
                            item.get("url")])
        for item in detail.get("economic_events", []):
            key = ("event", str(item.get("event")), str(item.get("scheduled_at")))
            if key in seen:
                continue
            seen.add(key)
            _append(sheet, [row["id"], _time(row.get("created_at")), row.get("symbol"),
                            "CALENDÁRIO", _time(item.get("scheduled_at")), item.get("event"),
                            item.get("currency"), item.get("impact"),
                            str(item.get("impact", "")).upper() in {"HIGH", "3", "3.0"}, ""])
    _finish(sheet)
    workbook.save(output)
    workbook.close()
    return output
